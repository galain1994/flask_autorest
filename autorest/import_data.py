# -*- coding: utf-8 -*-


import mimetypes
from collections import defaultdict
from sqlalchemy.inspection import inspect as sqla_inspect
from sqlalchemy.orm.base import MANYTOONE
from openpyxl import load_workbook


class Node(object):

    def __init__(self, name='root', parent=None, default_attrs=None, children=None, line=None):
        self.name = name
        self.parent = parent
        self.attributes = default_attrs or {}
        self.children = children or defaultdict(list)
        # 所有节点所在的行(包括明细行)
        self.lines = {line, } if line else set()

    def combine_children(self):
        """所有children都合并"""
        for relation, children in self.children.items():
            # record index
            rec_index = 0
            index_to_remove = []
            for i, node in enumerate(children):
                if not node.attributes and node.children and 0 != i:
                    self.children[relation][rec_index].combine_node(node)
                    index_to_remove.append(i)
                elif node and node.attributes:
                    rec_index = i
            for j in reversed(sorted(index_to_remove)):
                self.children[relation].pop(j)

    def combine_node(self, other):
        for key in set(self.children.keys()).union(set(other.children.keys())):
            self.children[key].extend(other.children[key])
            self.lines.update(other.lines)

    def add_attr(self, key, value, line):
        """设置当前node的属性和子行"""
        if self.attributes.get(key):
            return
        if '__' in key:
            relation, descendants = key.split('__', 1)
            if relation not in self.children:
                child = Node(relation, line=line)
                self.add_child(child)
            child = self.children[relation][-1]
            child.add_attr(descendants, value, line)
        else:
            self.attributes[key] = value

    def add_lines(self, node):
        if node.lines:
            self.lines.update(node.lines)

    def add_child(self, node):
        self.children[node.name].append(node)
        self.add_lines(node)
        node.parent = self

    def present(self):
        """Convert node to dictionary"""
        data = {k: v for k, v in self.attributes.items()}
        for relation, children in self.children.items():
            data[relation] = [child.present() for child in children]
        return data

    def deserialize(self, model):
        """将行"""
        inspected = sqla_inspect(model)
        drop_relations = []
        for relation, children in self.children.items():
            if relation not in inspected.relationships:
                # 不属于关联模型的数据
                drop_relations.append(relation)
                continue
            _model = inspected.relationships[relation].mapper.class_
            direction = inspected.relationships[relation].direction
            if MANYTOONE == direction:
                child = children[0]
                child.deserialize(_model)
                self.children[relation] = child
            else:
                for child in children:
                    child.deserialize(_model)
        for relation in drop_relations:
            self.children.pop(relation)
        return self.present()


def guess_file_format(upload_file):
    """根据请求头/文件开头猜测文件类型"""
    content_type = upload_file.mimetype
    if not content_type:
        content_type, encoding = mimetypes.guess_type(upload_file.filename)
    if 'text/csv' == content_type.strip().lower():
        return 'csv'
    elif content_type.strip().lower() in (
            'application/vnd.ms-excel',
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'):
        return 'excel'
    return None


def load_upload_file(upload_file, fmt):
    if 'csv' == fmt:
        return load_csv(upload_file)
    elif 'excel' == fmt:
        return load_excel(upload_file)
    raise ValueError("upload_file's format should be csv/excel")


def load_excel(upload_file):
    """
    加载上传的csv文件
    每行转化为一个list
    """
    workbook = load_workbook(upload_file)
    sheetname = workbook.sheetnames[0]
    sheet = workbook[sheetname]
    return [
        [cell.value or '' for cell in row]
        for row in sheet.rows
    ]


def load_csv(upload_file):
    """
    加载上传的csv文件
    每行转化为一个list
    """
    return [
        line.decode('utf-8').strip().split(',')
        for line in upload_file.readlines()
    ]


def import2lines(import_data):
    """ Data include headers
    return lines
    >>> imported = [['a', 'b__b1', 'b__b2', 'c', 'c__c1', 'c__c2'], ['*', '-', '-', '', '#', '#'], ['', '', '', '', '+', '+'], ['**', '--', '--', None, '', '']]
    >>> import2lines(imported)
    [{'a': '*', 'b__b1': '-', 'b__b2': '-', 'c__c1': '#', 'c__c2': '#'},
    {'c__c1': '+', 'c__c2': '+'},
    {'a': '**', 'b__b1': '--', 'b__b2': '--'}]
    """
    return [
        {head: row[i] for i, head in enumerate(import_data[0]) if row[i]}
        for row in import_data[1:]
    ]


def lines2root(lines, table_name='record'):
    """Change to tree structure and combine same root"""
    root = Node()
    for i, line_data in enumerate(lines):
        row = i + 1
        node = Node(table_name, line=row)
        for k, v in line_data.items():
            node.add_attr(k, v, row)
        root.add_child(node)
    root.combine_children()
    return root


def deserialize_root(root, model):
    return {
        ','.join([str(i) for i in child.lines]): child.deserialize(model)
        for child in list(root.children.values())[0]
    }
